from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .metric_calculator import aggregate_metrics
from .rule_engine import executive_summary
from .utils import ROOT


EXPORT_DIR = ROOT / "data" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _safe_sheet(df: pd.DataFrame, max_rows: int | None = None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame({"提示": ["暂无数据"]})
    return df.head(max_rows) if max_rows else df


def generate_excel_report(
    cleaned: pd.DataFrame,
    suggestions: pd.DataFrame,
    settings: dict,
    quality_report: pd.DataFrame | None = None,
    output_path: str | Path | None = None,
) -> Path:
    date_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path(output_path or EXPORT_DIR / f"Amazon_Ads_Optimization_Report_{date_tag}.xlsx")
    account = aggregate_metrics(cleaned, []) if not cleaned.empty else pd.DataFrame()
    campaign = aggregate_metrics(cleaned, ["campaign"]) if not cleaned.empty else pd.DataFrame()
    ad_group = aggregate_metrics(cleaned, ["campaign", "ad_group"]) if not cleaned.empty else pd.DataFrame()
    targeting = aggregate_metrics(cleaned, ["campaign", "ad_group", "targeting"]) if "targeting" in cleaned else pd.DataFrame()
    search = aggregate_metrics(cleaned, ["campaign", "ad_group", "search_term"]) if "search_term" in cleaned else pd.DataFrame()
    sku = aggregate_metrics(cleaned, ["sku", "asin"]) if "sku" in cleaned or "asin" in cleaned else pd.DataFrame()
    summary = executive_summary(cleaned, suggestions, settings)
    bid = suggestions[suggestions["action_type"].isin(["Increase Bid", "Decrease Bid"])] if not suggestions.empty else pd.DataFrame()
    budget = suggestions[suggestions["action_type"].isin(["Increase Budget", "Decrease Budget"])] if not suggestions.empty else pd.DataFrame()
    negative = suggestions[suggestions["action_type"].isin(["Add Negative Exact", "Add Negative Phrase"])] if not suggestions.empty else pd.DataFrame()
    harvest = suggestions[suggestions["action_type"].eq("Harvest Search Term")] if not suggestions.empty else pd.DataFrame()
    placement = aggregate_metrics(cleaned, ["placement"]) if "placement" in cleaned else pd.DataFrame()
    profit = aggregate_metrics(cleaned, ["campaign"]) if "campaign" in cleaned else pd.DataFrame()
    rule_log = suggestions.copy()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"管理层总结": summary.split("\n")}).to_excel(writer, sheet_name="Executive Summary", index=False)
        _safe_sheet(account).to_excel(writer, sheet_name="Account Overview", index=False)
        _safe_sheet(campaign).to_excel(writer, sheet_name="Campaign Analysis", index=False)
        _safe_sheet(ad_group).to_excel(writer, sheet_name="Ad Group Analysis", index=False)
        _safe_sheet(targeting).to_excel(writer, sheet_name="Targeting Analysis", index=False)
        _safe_sheet(search).to_excel(writer, sheet_name="Search Term Analysis", index=False)
        _safe_sheet(bid).to_excel(writer, sheet_name="Bid Recommendations", index=False)
        _safe_sheet(budget).to_excel(writer, sheet_name="Budget Recommendations", index=False)
        _safe_sheet(negative).to_excel(writer, sheet_name="Negative Keyword Suggestions", index=False)
        _safe_sheet(harvest).to_excel(writer, sheet_name="Search Term Harvesting", index=False)
        _safe_sheet(placement).to_excel(writer, sheet_name="Placement Analysis", index=False)
        _safe_sheet(sku).to_excel(writer, sheet_name="SKU Analysis", index=False)
        _safe_sheet(profit).to_excel(writer, sheet_name="Profit Analysis", index=False)
        _safe_sheet(quality_report if quality_report is not None else pd.DataFrame()).to_excel(writer, sheet_name="Data Quality Report", index=False)
        _safe_sheet(rule_log).to_excel(writer, sheet_name="Rule Trigger Log", index=False)
        _safe_sheet(cleaned, max_rows=100000).to_excel(writer, sheet_name="Raw Data Cleaned", index=False)
    format_workbook(path)
    return path


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill_p0 = PatternFill("solid", fgColor="FFC7CE")
    fill_p1 = PatternFill("solid", fgColor="FFEB9C")
    fill_p2 = PatternFill("solid", fgColor="D9EAD3")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    percent_names = {"ctr", "cvr", "acos", "tacos", "advertising_sales_share", "advertising_order_share", "bid_change_pct", "break_even_acos"}
    money_names = {"spend", "sales", "cpc", "cpa", "rpc", "bid", "budget", "current_bid", "recommended_bid", "bid_change", "target_cpc", "break_even_cpc", "profit_after_ads"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        headers = {cell.column: str(cell.value or "").lower() for cell in ws[1]}
        for col_idx, header in headers.items():
            width = min(max(len(header) + 4, 12), 45)
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = width
            for cell in ws[letter][1:]:
                if header in percent_names:
                    cell.number_format = "0.0%"
                elif header in money_names:
                    cell.number_format = '"$"#,##0.00'
                if header == "priority":
                    if cell.value == "P0":
                        cell.fill = fill_p0
                    elif cell.value == "P1":
                        cell.fill = fill_p1
                    elif cell.value == "P2":
                        cell.fill = fill_p2
    wb.save(path)
